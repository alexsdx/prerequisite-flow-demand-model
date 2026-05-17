"""Generate Validation_and_Trends.xlsx from the anonymised data.

The workbook summarises the validation evidence reported in Sections 5
and 7 of the article. It contains:

  Sheet 1  - Headline metrics (Experiment 1 and Experiment 2).
  Sheet 2  - Per-course validation table for Experiment 1 (S4 -> S5).
  Sheet 3  - Per-course validation table for Experiment 2 (S5 -> S6).
  Sheet 4  - Six-semester enrolment matrix (Tabla 1 of the article).
  Sheet 5  - Prerequisite graph as an edge list.

Unlike the development workbook used during analysis, this script reads
only the anonymised CSV and JSON files in data/. It produces results
that are bit-identical to those reported in the article.
"""

from __future__ import annotations

import csv
import json
import math
from pathlib import Path

from algorithm1 import compute_prerequisite_flow_demand

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Install with: pip install -r requirements.txt"
    ) from exc

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
OUT_PATH = REPO_ROOT / "Validation_and_Trends.xlsx"

SEMESTERS = ["S1", "S2", "S3", "S4", "S5", "S6"]
TAU_DEFAULT = 18

# --------------------------------------------------------------------------
# I/O
# --------------------------------------------------------------------------

def _load_enrolments() -> dict:
    rows = []
    with (DATA_DIR / "enrollments_anonymized.csv").open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append((row["course_id"], row["semester"], int(row["enrolment_n"])))
    table = {(c, s): n for c, s, n in rows}
    return table


def _load_failures_S4() -> dict:
    table = {}
    with (DATA_DIR / "failures_S4.csv").open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            table[(row["course_id"], "S4")] = int(row["failures_f"])
    return table


def _load_graph() -> dict:
    with (DATA_DIR / "prerequisite_graph.json").open(encoding="utf-8") as f:
        return json.load(f)


# --------------------------------------------------------------------------
# Styles
# --------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_TITLE_FONT = Font(bold=True, size=14, color="2F5496")
_NOTE_FONT = Font(italic=True, size=9, color="595959")
_BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_header(ws, row, columns):
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _write_row(ws, row, values, align_left_cols=()):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = _BORDER
        cell.alignment = _LEFT if col_idx in align_left_cols else _CENTER


def _set_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# --------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------

def sheet_headline(wb, exp1_summary, exp2_summary):
    ws = wb.active
    ws.title = "1. Headline Metrics"
    ws.cell(row=1, column=1, value="HEADLINE VALIDATION METRICS").font = _TITLE_FONT
    ws.cell(row=2, column=1, value="Sections 5 of the article").font = _NOTE_FONT

    r = 4
    _write_header(ws, r, ["Metric", "Experiment 1 (S4 -> S5)", "Experiment 2 (S5 -> S6)"])
    r += 1

    rows = [
        ("Validated courses", len(exp1_summary["validated"]), len(exp2_summary["validated"])),
        ("Sum predicted demand", exp1_summary["sum_d"], exp2_summary["sum_d"]),
        ("Sum observed enrolment", exp1_summary["sum_n"], exp2_summary["sum_n"]),
        ("Coverage (%)", exp1_summary["coverage"], exp2_summary["coverage"]),
        ("MAE (students per course)", exp1_summary["mae"], exp2_summary["mae"]),
    ]
    for metric, v1, v2 in rows:
        _write_row(ws, r, [metric, v1, v2], align_left_cols=(1,))
        r += 1

    _set_widths(ws, [34, 26, 26])


def sheet_per_course(wb, title, predictions, enrolments, target_semester, tab_color):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT

    r = 3
    _write_header(ws, r, [
        "Course", "Predicted d_c", "Observed n_c",
        "Residual (n - d)", "Sign", "Groups @ tau=18",
    ])
    r += 1

    for course in sorted(predictions):
        d = predictions[course]
        if d <= 0:
            continue
        n = enrolments.get((course, target_semester), 0)
        if n <= 0:
            continue
        residual = n - d
        sign = "+" if residual > 0 else ("-" if residual < 0 else "0")
        groups = math.ceil(d / TAU_DEFAULT)
        _write_row(ws, r, [course, d, n, residual, sign, groups],
                   align_left_cols=(1,))
        r += 1

    _set_widths(ws, [14, 16, 16, 16, 8, 18])


def sheet_enrolment_matrix(wb, enrolments, graph):
    ws = wb.create_sheet("4. Enrolment Matrix")
    ws.cell(row=1, column=1, value="SIX-SEMESTER ENROLMENT MATRIX").font = _TITLE_FONT
    ws.cell(row=2, column=1, value="Table 1 of the article (with anonymised labels).").font = _NOTE_FONT

    r = 4
    _write_header(ws, r, ["Course"] + SEMESTERS)
    r += 1

    for course in graph["vertices"]:
        row_values = [course]
        any_enrolled = False
        for sem in SEMESTERS:
            n = enrolments.get((course, sem), 0)
            row_values.append(n if n > 0 else "-")
            if n > 0:
                any_enrolled = True
        if not any_enrolled:
            continue
        _write_row(ws, r, row_values, align_left_cols=(1,))
        r += 1

    _set_widths(ws, [12] + [12] * len(SEMESTERS))


def sheet_graph(wb, graph):
    ws = wb.create_sheet("5. Prerequisite Graph")
    ws.cell(row=1, column=1, value="PREREQUISITE GRAPH (EDGE LIST)").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "Forest of in-trees; each course has at most one prerequisite."
    )).font = _NOTE_FONT

    r = 4
    _write_header(ws, r, ["Prerequisite", "Course"])
    r += 1
    for edge in graph["edges"]:
        _write_row(ws, r, [edge["from"], edge["to"]], align_left_cols=(1, 2))
        r += 1

    _set_widths(ws, [16, 16])


# --------------------------------------------------------------------------
# Driver
# --------------------------------------------------------------------------

def _evaluate(predictions, enrolments, target_semester):
    validated = []
    sum_d = sum_n = err_total = 0
    for course, d in predictions.items():
        if d <= 0:
            continue
        n = enrolments.get((course, target_semester), 0)
        if n <= 0:
            continue
        validated.append(course)
        sum_d += d
        sum_n += n
        err_total += abs(d - n)
    coverage = round(sum_d / sum_n * 100, 1) if sum_n else 0.0
    mae = round(err_total / len(validated), 1) if validated else 0.0
    return {
        "validated": validated, "sum_d": sum_d, "sum_n": sum_n,
        "coverage": coverage, "mae": mae,
    }


def main():
    enrolments = _load_enrolments()
    failures_S4 = _load_failures_S4()
    graph = _load_graph()

    pred_exp1 = compute_prerequisite_flow_demand(
        graph=graph, enrolments=enrolments, failures=failures_S4,
        t="S4", t_plus_1="S5", semester_order=SEMESTERS, k=1,
    )
    pred_exp2 = compute_prerequisite_flow_demand(
        graph=graph, enrolments=enrolments, failures={},
        t="S5", t_plus_1="S6", semester_order=SEMESTERS, k=1,
    )

    summary1 = _evaluate(pred_exp1, enrolments, "S5")
    summary2 = _evaluate(pred_exp2, enrolments, "S6")

    wb = Workbook()
    sheet_headline(wb, summary1, summary2)
    sheet_per_course(
        wb, "2. Experiment 1 (S4 to S5)", pred_exp1, enrolments,
        target_semester="S5", tab_color="2F5496",
    )
    sheet_per_course(
        wb, "3. Experiment 2 (S5 to S6)", pred_exp2, enrolments,
        target_semester="S6", tab_color="00B050",
    )
    sheet_enrolment_matrix(wb, enrolments, graph)
    sheet_graph(wb, graph)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH.name}")


if __name__ == "__main__":
    main()
